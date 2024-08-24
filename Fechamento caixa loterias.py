import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

class FechamentoCaixaApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Fechamento de Caixa de Loteria")

        # Definindo as categorias de entrada
        self.categorias = ["Vale",
            "+ DH", "+ Despesas (Loteria)", "+ Despesas (Lúcio)", "+ Energia/Água/Aluguel",
            "+ Resgate", "+ Diversos", "+ Jogos Clientes", "+ Vale Funcionário",
            "+ Fundo de Caixa", "+ Caixa (almoço)", "- Total em Caixa (Fita)",
            "- Tele Sena / Federal", "- Lucro Bolão (35%)", "- Tarifa Cheque",
            "- Reforço Caixa", "+ Depósito (d+1)", "+ Depósitos Realizados", "+ Depósito (d-1)", "+ Produto (Dia Anterior)"
        ]

        # Lista vazia para armazenar os nomes dos funcionários
        self.funcionarios = []

        # Variável para armazenar o nome da loteria
        self.nome_loteria = ""

        # Dicionário para armazenar as entradas dos valores
        self.dados = {}

        # Caixa de entrada para os nomes dos funcionários e nome da loteria
        self.criar_caixa_nomes_funcionarios()

    def criar_caixa_nomes_funcionarios(self):
        # Rótulo e entrada para solicitar os nomes dos funcionários
        self.label_instrucao = tk.Label(self.root, text="Digite os nomes dos funcionários separados por vírgula:")
        self.label_instrucao.grid(row=0, column=0, columnspan=len(self.categorias)+2, padx=10, pady=10)

        self.entry_nomes_funcionarios = tk.Entry(self.root, width=50)
        self.entry_nomes_funcionarios.grid(row=1, column=0, columnspan=len(self.categorias)+2, padx=10, pady=5)

        # Rótulo e entrada para solicitar o nome da loteria
        self.label_nome_loteria = tk.Label(self.root, text="Digite o nome da loteria:")
        self.label_nome_loteria.grid(row=2, column=0, padx=10, pady=5)

        self.entry_nome_loteria = tk.Entry(self.root, width=50)
        self.entry_nome_loteria.grid(row=3, column=0, columnspan=len(self.categorias)+2, padx=10, pady=5)

        self.btn_confirmar_nomes = tk.Button(self.root, text="Confirmar Nomes e Loteria", command=self.confirmar_nomes_funcionarios)
        self.btn_confirmar_nomes.grid(row=4, column=0, columnspan=len(self.categorias)+2, pady=10)

    def confirmar_nomes_funcionarios(self):
        # Obtendo os nomes digitados pelo usuário
        nomes = self.entry_nomes_funcionarios.get().strip()
        self.funcionarios = [nome.strip() for nome in nomes.split(',') if nome.strip()]

        # Obtendo o nome da loteria
        self.nome_loteria = self.entry_nome_loteria.get().strip()

        # Verificando se foram fornecidos pelo menos 1 nome e um nome de loteria
        if len(self.funcionarios) < 1 or not self.nome_loteria:
            tk.messagebox.showerror("Erro", "Por favor, insira pelo menos um nome de funcionário e o nome da loteria.")
            return

        # Remover a caixa de entrada e o botão após confirmar os nomes
        self.label_instrucao.destroy()
        self.entry_nomes_funcionarios.destroy()
        self.label_nome_loteria.destroy()
        self.entry_nome_loteria.destroy()
        self.btn_confirmar_nomes.destroy()

        # Criar a grade para entrada de valores
        self.criar_widgets()

    def criar_widgets(self):
        # Rótulos dos nomes dos funcionários (na linha 2)
        for j, funcionario in enumerate(self.funcionarios):
            label_funcionario = tk.Label(self.root, text=funcionario)
            label_funcionario.grid(row=2, column=j+2, padx=10, pady=5, sticky='w')

        # Criação da grade para entrada de valores
        for i, categoria in enumerate(self.categorias):
            # Rótulo da categoria
            label_categoria = tk.Label(self.root, text=categoria)
            label_categoria.grid(row=i+3, column=0, padx=10, pady=5, sticky='w')

            if categoria in ["+ Depósito (d+1)", "+ Depósito (d-1)", "+ Depósitos Realizados", "+ Produto (Dia Anterior)"]:
                # Para essas categorias, criar uma entrada para cada funcionário
                for j, funcionario in enumerate(self.funcionarios):
                    # Entrada para digitar o valor
                    entry_valor = tk.Entry(self.root, width=10)
                    entry_valor.grid(row=i+3, column=j+2, padx=10, pady=5)  # Entrada na linha i+3, coluna j+2

                    # Armazenar a entrada no dicionário de dados
                    if categoria not in self.dados:
                        self.dados[categoria] = {}
                    self.dados[categoria][funcionario] = entry_valor

            else:
                # Para as demais categorias, criar uma entrada para cada funcionário
                for j, funcionario in enumerate(self.funcionarios):
                    entry_valor = tk.Entry(self.root, width=10)
                    entry_valor.grid(row=i+3, column=j+2, padx=10, pady=5)  # Entrada na linha i+3, coluna j+2
                    if categoria not in self.dados:
                        self.dados[categoria] = {}
                    self.dados[categoria][funcionario] = entry_valor

        # Botão para calcular a diferença do caixa
        btn_calcular = tk.Button(self.root, text="Calcular Fechamento do Caixa", command=self.calcular_fechamento_caixa)
        btn_calcular.grid(row=len(self.categorias)+4, column=0, columnspan=len(self.funcionarios)+2, pady=10, padx=10)  # Botão alinhado corretamente

        # Label para exibir o resultado do fechamento do caixa
        self.label_fechamento_caixa = tk.Label(self.root, text="", justify='left')
        self.label_fechamento_caixa.grid(row=3, rowspan=len(self.categorias), column=len(self.funcionarios)+2, padx=10, pady=10, sticky='n')  # Label alinhado corretamente

    def calcular_fechamento_caixa(self):
        total_itens_caixa = {funcionario: 0 for funcionario in self.funcionarios}
        total_outros_itens = {funcionario: 0 for funcionario in self.funcionarios}
        total_depositos = 0

        # Obtendo os valores inseridos pelos funcionários
        for categoria in self.categorias:
            for funcionario in self.funcionarios:
                valor = float(self.dados[categoria][funcionario].get() or 0)
                if categoria in ["Vale","+ DH", "+ Despesas (Loteria)", "+ Despesas (Lúcio)", "+ Energia/Água/Aluguel",
                                 "+ Resgate", "+ Diversos", "+ Jogos Clientes", "+ Vale Funcionário",
                                 "+ Fundo de Caixa", "+ Caixa (almoço)"]:
                    total_itens_caixa[funcionario] += valor
                elif categoria in ["- Total em Caixa (Fita)", "- Tele Sena / Federal", "- Lucro Bolão (35%)",
                                   "- Tarifa Cheque", "- Reforço Caixa"]:
                    total_outros_itens[funcionario] += valor

                # Calcular depósito
                if categoria in ["+ Depósito (d+1)", "+ Depósitos Realizados"]:
                    total_depositos += valor

        # Calculando o fechamento do caixa para cada funcionário
        fechamento_caixa = {funcionario: total_itens_caixa[funcionario] - total_outros_itens[funcionario] for funcionario in self.funcionarios}

        # Calculando a variável "+ Miudezas"
        soma_miudezas = 0
        for categoria in ["+ Despesas (Loteria)", "+ Despesas (Lúcio)", "+ Energia/Água/Aluguel",
                        "+ Resgate", "+ Diversos", "+ Jogos Clientes", "+ Vale Funcionário",
                        "+ Fundo de Caixa"]:
            for funcionario in self.funcionarios:
                soma_miudezas += float(self.dados[categoria][funcionario].get() or 0)

        # Subtraindo a soma dos fechamentos de caixa de todos os funcionários
        soma_fechamentos = sum(fechamento_caixa.values())
        miudezas = soma_miudezas - soma_fechamentos

        # Calculando a variável "- Fitas + Lucro Bolão + Tele"
        soma_fitas_lucro_bolao_tele = 0
        valor_total_em_caixa_fita_ultimo_funcionario = 0

        for i, categoria in enumerate(self.categorias):
            if categoria == "- Total em Caixa (Fita)":
                for j in range(i+1, len(self.categorias)):
                    if self.categorias[j] == "- Reforço Caixa":
                        # Encontrar o valor de "- Total em Caixa (Fita)" do último funcionário
                        valor_total_em_caixa_fita_ultimo_funcionario = float(self.dados[categoria][self.funcionarios[-1]].get() or 0)
                        break

        # Soma todas as categorias entre "- Total em Caixa (Fita)" e "- Reforço Caixa"
        for categoria in ["- Total em Caixa (Fita)",
            "- Tele Sena / Federal", "- Lucro Bolão (35%)", "- Tarifa Cheque",
            "- Reforço Caixa"]:
            for funcionario in self.funcionarios:
                soma_fitas_lucro_bolao_tele += float(self.dados[categoria][funcionario].get() or 0)

        # Subtrai o valor de "- Total em Caixa (Fita)" do último funcionário
        fitas_lucro_bolao_tele = soma_fitas_lucro_bolao_tele - valor_total_em_caixa_fita_ultimo_funcionario

        # Calculando a variável "+ Suplemento"  
        suplemento = sum(float(self.dados["- Reforço Caixa"][funcionario].get() or 0) for funcionario in self.funcionarios)

        # Calculando a variável "- Depósito D-1"
        deposito_d_1 = sum(float(self.dados["+ Depósito (d-1)"][funcionario].get() or 0) for funcionario in self.funcionarios)

        # Calculando a variável "- Produtos (dia anterior)"
        produtos_dia_anterior = sum(float(self.dados["+ Produto (Dia Anterior)"][funcionario].get() or 0) for funcionario in self.funcionarios)

        # Calculando o resultado final conforme solicitado
        resultado_final = miudezas + total_depositos + suplemento - fitas_lucro_bolao_tele - deposito_d_1 - produtos_dia_anterior

        # Exibindo o resultado do fechamento do caixa na interface
        texto_resultado = "\n".join([f"{funcionario}: R$ {fechamento_caixa[funcionario]:.2f}" for funcionario in self.funcionarios])
        texto_resultado += f"\n\nDepósitos Realizados: R$ {total_depositos:.2f}"
        texto_resultado += f"\n+ Miudezas: R$ {miudezas:.2f}"  # Adicionando a variável "+ Miudezas"
        texto_resultado += f"\n- Fitas + Lucro Bolão + Tele: R$ {fitas_lucro_bolao_tele:.2f}"  # Adicionando a variável "- Fitas + Lucro Bolão + Tele"
        texto_resultado += f"\n+ Suplemento: R$ {suplemento:.2f}"  # Adicionando a variável "+ Suplemento"
        texto_resultado += f"\n- Depósito D-1: R$ {deposito_d_1:.2f}"  # Adicionando a variável "- Depósito D-1"
        texto_resultado += f"\n- Produtos (dia anterior): R$ {produtos_dia_anterior:.2f}"  # Adicionando a variável "- Produtos (dia anterior)"
        texto_resultado += f"\n\nResultado Final: R$ {resultado_final:.2f}"
        self.label_fechamento_caixa.config(text=f"Fechamento do Caixa:\n{texto_resultado}")

        # Salvar dados em arquivo Excel
        self.salvar_em_excel()

        # Não abrir mais a janela de edição de depósitos
        self.finalizar()

    def salvar_em_excel(self):
        arquivo_excel = r"C:\Users\AGFAKZZ\Desktop\Caixa\Fechamento_Caixa.xlsx"
        cabeçalhos = ["Nome da Loteria", "Funcionário"] + self.categorias + ["Data_Hora"]

        if os.path.exists(arquivo_excel):
            # Se o arquivo já existe, carregue-o
            wb = load_workbook(arquivo_excel)
            ws = wb.active
        else:
            # Se o arquivo não existe, crie um novo
            wb = Workbook()
            ws = wb.active
            ws.title = "Fechamento de Caixa"
            # Adiciona os cabeçalhos
            ws.append(cabeçalhos)

        # Adicionar os dados para cada funcionário
        for funcionario in self.funcionarios:
            row = [self.nome_loteria, funcionario]
            for categoria in self.categorias:
                valor = self.dados[categoria].get(funcionario, tk.Entry(self.root)).get() or '0'
                row.append(float(valor))
            # Adiciona a data e hora atual
            row.append(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            ws.append(row)

        # Salvar o arquivo Excel
        wb.save(arquivo_excel)
        tk.messagebox.showinfo("Salvamento", f"Dados salvos em {arquivo_excel}")

# Função principal para inicializar a aplicação
if __name__ == "__main__":
    root = tk.Tk()
    root.geometry("900x500")  # Definindo a largura e altura da janela
    app = FechamentoCaixaApp(root)
    root.mainloop()